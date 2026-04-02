"""Client KoboCollect API et générateur XLSForm pour AgriTogo."""

import csv
import io
import json
import os
import requests


CONFIG_PATH = os.path.join(os.path.dirname(__file__), "kobo_config.json")


class KoboClient:
    """Client pour l'API KoboToolbox v2."""

    def __init__(self, base_url: str, token: str):
        self.base_url = base_url.rstrip("/")
        self.token = token
        self.headers = {"Authorization": f"Token {token}"}

    def get_forms(self) -> list:
        """Récupère la liste des formulaires."""
        try:
            r = requests.get(
                f"{self.base_url}/api/v2/assets/?format=json",
                headers=self.headers, timeout=15,
            )
            r.raise_for_status()
            return [
                {"uid": a["uid"], "name": a["name"],
                 "deployment_status": a.get("deployment_status", "")}
                for a in r.json().get("results", [])
            ]
        except Exception:
            return []

    def get_submissions(self, form_uid: str) -> list:
        """Récupère les soumissions d'un formulaire."""
        try:
            r = requests.get(
                f"{self.base_url}/api/v2/assets/{form_uid}/data/?format=json",
                headers=self.headers, timeout=15,
            )
            r.raise_for_status()
            return r.json().get("results", [])
        except Exception:
            return []

    def get_form_count(self) -> int:
        """Retourne le nombre total de formulaires."""
        return len(self.get_forms())


def generate_price_survey_xlsform() -> dict:
    """Génère un XLSForm pour la collecte de prix sur les marchés togolais."""
    markets = ["Lome-Adawlato", "Kara", "Sokode", "Atakpame", "Dapaong"]
    products = ["Mais", "Riz", "Sorgho", "Mil", "Haricot", "Soja",
                "Arachide", "Igname", "Manioc", "Tomate", "Piment", "Oignon"]
    units = ["kg", "tonne", "sac"]
    def _f(t, n, l, req="yes", c=""):
        return {"type": t, "name": n, "label": l, "required": req, "constraint": c}
    survey = [
        _f("today", "date", "Date de collecte"),
        _f("select_one market", "market", "Marché"),
        _f("select_one product", "product", "Produit"),
        _f("integer", "price", "Prix (FCFA)", c=". > 0"),
        _f("select_one unit", "unit", "Unité"),
        _f("text", "collector_name", "Nom du collecteur"),
        _f("geopoint", "gps_location", "Position GPS", req="no"),
        _f("text", "notes", "Notes", req="no"),
    ]
    choices = (
        [{"list_name": "market", "name": m, "label": m} for m in markets]
        + [{"list_name": "product", "name": p, "label": p} for p in products]
        + [{"list_name": "unit", "name": u, "label": u} for u in units]
    )
    return {"survey": survey, "choices": choices,
            "title": "AgriTogo - Collecte Prix Marches", "form_id": "agritogo_prix"}


def generate_farmer_survey_xlsform() -> dict:
    """Génère un XLSForm pour le profilage des agriculteurs togolais."""
    regions = ["Maritime", "Plateaux", "Centrale", "Kara", "Savanes"]
    crops = ["Mais", "Riz", "Sorgho", "Mil", "Igname", "Manioc",
             "Soja", "Arachide", "Tomate", "Piment"]
    yn = ["yes", "no"]
    storage = ["none", "basic", "good"]
    def _f(t, n, l, req="yes", c=""):
        return {"type": t, "name": n, "label": l, "required": req, "constraint": c}
    survey = [
        _f("text", "farmer_name", "Nom de l'agriculteur"),
        _f("select_one region", "region", "Région"),
        _f("decimal", "farm_size_ha", "Superficie (ha)", c=". > 0"),
        _f("select_one crop", "main_crop", "Culture principale"),
        _f("select_one crop", "secondary_crop", "Culture secondaire", req="no"),
        _f("integer", "annual_revenue_fcfa", "Revenu annuel (FCFA)", req="no", c=". >= 0"),
        _f("select_one yn", "has_irrigation", "Irrigation ?"),
        _f("select_one yn", "has_insurance", "Assurance ?"),
        _f("integer", "years_experience", "Années d'expérience", c=". >= 0"),
        _f("select_one storage", "storage_capacity", "Capacité de stockage"),
        _f("decimal", "distance_to_market_km", "Distance au marché (km)", req="no", c=". >= 0"),
        _f("text", "notes", "Notes", req="no"),
    ]
    choices = (
        [{"list_name": "region", "name": r, "label": r} for r in regions]
        + [{"list_name": "crop", "name": c, "label": c} for c in crops]
        + [{"list_name": "yn", "name": v, "label": v} for v in yn]
        + [{"list_name": "storage", "name": s, "label": s} for s in storage]
    )
    return {"survey": survey, "choices": choices,
            "title": "AgriTogo - Profil Agriculteur", "form_id": "agritogo_farmer"}


def generate_crop_yield_form() -> dict:
    """Génère un XLSForm pour la collecte de données de rendement (module ML crop_yield)."""
    regions = ["Maritime", "Plateaux", "Centrale", "Kara", "Savanes"]
    crops = ["Mais", "Riz", "Sorgho", "Mil", "Igname", "Manioc", "Soja", "Arachide"]
    soil_health = ["poor", "average", "good", "excellent"]
    yn = ["yes", "no"]
    seasons = ["grande_saison", "petite_saison", "saison_seche"]

    def _f(t, n, l, req="yes", c=""):
        return {"type": t, "name": n, "label": l, "required": req, "constraint": c}

    survey = [
        _f("text", "farmer_id", "Identifiant agriculteur"),
        _f("select_one region", "region", "Région"),
        _f("select_one crop", "crop_type", "Culture principale"),
        _f("decimal", "farm_size_ha", "Superficie cultivée (ha)", c=". > 0"),
        _f("decimal", "yield_kg_ha", "Rendement obtenu (kg/ha)", c=". > 0"),
        _f("decimal", "avg_temperature_c", "Température moyenne observée (°C)"),
        _f("decimal", "rainfall_mm", "Pluviométrie estimée (mm)"),
        _f("decimal", "pesticides_kg_ha", "Pesticides utilisés (kg/ha)"),
        _f("decimal", "fertilizer_kg_ha", "Engrais utilisés (kg/ha)"),
        _f("select_one soil_health", "soil_health_score", "Qualité du sol"),
        _f("integer", "extreme_weather_events", "Événements météo extrêmes (nombre)"),
        _f("select_one yn", "irrigation_access", "Accès à l'irrigation"),
        _f("select_one season", "season", "Saison de culture"),
        _f("geopoint", "gps_location", "Position GPS", req="no"),
        _f("text", "notes", "Observations", req="no"),
    ]
    choices = (
        [{"list_name": "region", "name": r, "label": r} for r in regions]
        + [{"list_name": "crop", "name": c, "label": c} for c in crops]
        + [{"list_name": "soil_health", "name": s, "label": s} for s in soil_health]
        + [{"list_name": "yn", "name": v, "label": v} for v in yn]
        + [{"list_name": "season", "name": s, "label": s} for s in seasons]
    )
    return {"survey": survey, "choices": choices,
            "title": "AgriTogo - Collecte Rendement", "form_id": "agritogo_yield"}


def generate_financial_risk_form() -> dict:
    """Génère un XLSForm pour l'évaluation du risque financier (module ML financial_risk)."""
    regions = ["Maritime", "Plateaux", "Centrale", "Kara", "Savanes"]
    enterprise_sizes = ["Small", "Medium", "Large"]
    yn = ["yes", "no"]
    risk_levels = ["low", "medium", "high"]
    crops = ["Mais", "Riz", "Sorgho", "Mil", "Igname", "Manioc", "Soja", "Arachide"]

    def _f(t, n, l, req="yes", c=""):
        return {"type": t, "name": n, "label": l, "required": req, "constraint": c}

    survey = [
        _f("text", "farmer_id", "Identifiant agriculteur"),
        _f("select_one region", "region", "Région"),
        _f("select_one enterprise_size", "enterprise_size", "Taille de l'exploitation"),
        _f("integer", "annual_revenue_fcfa", "Revenu annuel (FCFA)"),
        _f("integer", "annual_expenses_fcfa", "Dépenses annuelles (FCFA)"),
        _f("integer", "loan_amount_fcfa", "Montant du prêt (FCFA)"),
        _f("integer", "loan_duration_months", "Durée du prêt (mois)"),
        _f("select_one yn", "has_previous_default", "Défaut de paiement antérieur"),
        _f("decimal", "debt_to_equity_ratio", "Ratio dette/fonds propres"),
        _f("select_one risk_level", "drought_risk_perception", "Risque sécheresse perçu"),
        _f("select_one risk_level", "flood_risk_perception", "Risque inondation perçu"),
        _f("select_one yn", "has_insurance", "Assurance agricole"),
        _f("select_one crop", "main_crop", "Culture principale"),
        _f("integer", "years_farming", "Années d'expérience"),
        _f("geopoint", "gps_location", "Position GPS", req="no"),
    ]
    choices = (
        [{"list_name": "region", "name": r, "label": r} for r in regions]
        + [{"list_name": "enterprise_size", "name": s, "label": s} for s in enterprise_sizes]
        + [{"list_name": "yn", "name": v, "label": v} for v in yn]
        + [{"list_name": "risk_level", "name": r, "label": r} for r in risk_levels]
        + [{"list_name": "crop", "name": c, "label": c} for c in crops]
    )
    return {"survey": survey, "choices": choices,
            "title": "AgriTogo - Risque Financier", "form_id": "agritogo_risk"}


def generate_market_price_form() -> dict:
    """Génère un XLSForm amélioré pour la collecte de prix (module GARCH/graphiques prix)."""
    markets = ["Lome-Adawlato", "Kara", "Sokode", "Atakpame", "Dapaong"]
    products = ["Mais", "Riz_local", "Sorgho", "Mil", "Haricot", "Soja",
                "Arachide", "Igname", "Manioc", "Tomate", "Piment", "Oignon"]
    supply_levels = ["low", "medium", "high", "abundant"]
    demand_levels = ["low", "medium", "high", "very_high"]
    quality_levels = ["poor", "average", "good", "excellent"]
    origin_regions = ["Maritime", "Plateaux", "Centrale", "Kara", "Savanes", "Import"]

    def _f(t, n, l, req="yes", c=""):
        return {"type": t, "name": n, "label": l, "required": req, "constraint": c}

    survey = [
        _f("text", "collector_id", "Identifiant collecteur"),
        _f("date", "collection_date", "Date de collecte"),
        _f("select_one market", "market", "Marché"),
        _f("select_one product", "product", "Produit"),
        _f("decimal", "price_fcfa_kg", "Prix (FCFA/kg)", c=". > 0"),
        _f("decimal", "price_min_fcfa", "Prix minimum observé"),
        _f("decimal", "price_max_fcfa", "Prix maximum observé"),
        _f("decimal", "volume_available_kg", "Volume disponible (kg)"),
        _f("select_one supply_level", "supply_level", "Niveau d'offre"),
        _f("select_one demand_level", "demand_level", "Niveau de demande"),
        _f("select_one quality", "quality", "Qualité du produit"),
        _f("select_one origin_region", "origin_region", "Région d'origine"),
        _f("decimal", "transport_cost_fcfa", "Coût transport (FCFA)", req="no"),
        _f("geopoint", "gps_location", "Position GPS", req="no"),
        _f("text", "notes", "Observations", req="no"),
    ]
    choices = (
        [{"list_name": "market", "name": m, "label": m} for m in markets]
        + [{"list_name": "product", "name": p, "label": p} for p in products]
        + [{"list_name": "supply_level", "name": s, "label": s} for s in supply_levels]
        + [{"list_name": "demand_level", "name": d, "label": d} for d in demand_levels]
        + [{"list_name": "quality", "name": q, "label": q} for q in quality_levels]
        + [{"list_name": "origin_region", "name": r, "label": r} for r in origin_regions]
    )
    return {"survey": survey, "choices": choices,
            "title": "AgriTogo - Prix Marché", "form_id": "agritogo_market"}


def xlsform_to_xlsx(xlsform_dict: dict) -> bytes:
    """Convertit un dict XLSForm en fichier Excel (.xlsx) avec feuilles survey, choices, settings."""
    from openpyxl import Workbook
    wb = Workbook()

    # Survey sheet
    ws_survey = wb.active
    ws_survey.title = "survey"
    if xlsform_dict.get("survey"):
        keys = list(xlsform_dict["survey"][0].keys())
        ws_survey.append(keys)
        for row in xlsform_dict["survey"]:
            ws_survey.append([row.get(k, "") for k in keys])

    # Choices sheet
    ws_choices = wb.create_sheet("choices")
    if xlsform_dict.get("choices"):
        keys = list(xlsform_dict["choices"][0].keys())
        ws_choices.append(keys)
        for row in xlsform_dict["choices"]:
            ws_choices.append([row.get(k, "") for k in keys])

    # Settings sheet
    ws_settings = wb.create_sheet("settings")
    ws_settings.append(["form_title", "form_id", "version"])
    title = xlsform_dict.get("title", "AgriTogo Survey")
    form_id = xlsform_dict.get("form_id", "agritogo_form")
    ws_settings.append([title, form_id, "1"])

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def save_kobo_config(base_url: str, token: str) -> None:
    """Sauvegarde la configuration KoboCollect."""
    with open(CONFIG_PATH, "w") as f:
        json.dump({"base_url": base_url, "token": token}, f)


def load_kobo_config() -> dict | None:
    """Charge la configuration KoboCollect, ou None si absente."""
    if not os.path.exists(CONFIG_PATH):
        return None
    with open(CONFIG_PATH) as f:
        return json.load(f)
